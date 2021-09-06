from ldap3 import Server, Connection, SUBTREE, ALL, NTLM
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl import Workbook
from datetime import datetime
import re,operator,pandas as pd,getpass

#Текущая дата
current_time = datetime.now()

#Настройки подключения к AD
server = Server('LDAP-DC.merlion.local', use_ssl=True, get_info=ALL)
domain = 'MERLION\\'

my_user = input('Введите пользователя: ')
my_secret = getpass.getpass('Введите пароль: ')

#Подключаемся к AD
conn = Connection(server, user=domain + my_user, password=my_secret, authentication=NTLM, return_empty_attributes=True, auto_bind=True)

#Поисковый запрос
entry_list = conn.extend.standard.paged_search(search_base = 'ou=users,ou=users_domain,dc=merlion,dc=local',search_filter ='(&(objectCategory=person)(objectClass=user) (!(userAccountControl:1.2.840.113556.1.4.803:=2)) (telephoneNumber=*) )',
                                               search_scope = SUBTREE,
                                               attributes=['extensionAttribute2','sAMAccountName','department','mail','telephoneNumber','Company','physicalDeliveryOfficeName','msRTCSIP-Line'],
                                               paged_size = 1000,
                                               generator=False)

#Закрываем соединение с AD
conn.unbind()

#Нормализуем данные
for res_tel in entry_list:
    all_tel = str(res_tel['attributes']['telephoneNumber']) + ',' +  str(res_tel['attributes']['msRTCSIP-Line'])
    clean_string = re.sub(r"[+()0-9 -]{10,20}", "", all_tel)
    iptel = re.findall(r'\d{4,5}', clean_string)
    # Получаем уникальные номера
    res_tel['iptel'] = set(iptel) if iptel else None

#Собираем свой список из словарей
persons_list = [
dict(user=item['attributes']['extensionAttribute2'],
     dept=item['attributes']['department'],
     tel=item['iptel'],
     login=item['attributes']['sAMAccountName'],
     email=item['attributes']['mail'],
     location=item['attributes']['Company'],
     office=item['attributes']['physicalDeliveryOfficeName']
     )

for item in entry_list
    if item['iptel']
]

#Один телефон одна запись
one_phone = pd.DataFrame(persons_list).explode("tel").to_dict("records")

#Сортируем словарь по ключу 'tel'
one_phone.sort(key=operator.itemgetter('tel'))

#Читаем файл "sp_net.txt"
with open('sp_net.txt', 'r') as f:
    busy_numbers = f.read().splitlines()

#Создаем файл Excel
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet()

#Называем лист Excel
ws.title = f"Users merlion.local {current_time.strftime('%d.%m.%Y')}"
ws1.title = "Занятые номера"

#Заморозка первой строки, столбца
wb[f"Users merlion.local {current_time.strftime('%d.%m.%Y')}"].freeze_panes = "B2"

#Счётчик ячеек
count = 2

#Задаем заголовки
ws['A1'] = 'Ф.И.О'
ws['B1'] = 'Отдел'
ws['C1'] = 'Внутренний телефон'
ws['D1'] = 'Учётная запись'
ws['E1'] = 'Электронная почта'
ws['F1'] = 'Местоположение'
ws['G1'] = 'Рабочее место'

#Задаем ширину столбцов
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 72
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 35
ws.column_dimensions["F"].width = 57
ws.column_dimensions["G"].width = 22

ws1.column_dimensions["A"].width = 60

#Выравниваем заголовок по центру
ws['A1'].alignment = Alignment(horizontal="center")
ws['B1'].alignment = Alignment(horizontal="center")
ws['C1'].alignment = Alignment(horizontal="center")
ws['D1'].alignment = Alignment(horizontal="center")
ws['E1'].alignment = Alignment(horizontal="center")
ws['F1'].alignment = Alignment(horizontal="center")
ws['G1'].alignment = Alignment(horizontal="center")

#Выделяем заголовок "жирным текстом"
ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['C1'].font = Font(bold=True)
ws['D1'].font = Font(bold=True)
ws['E1'].font = Font(bold=True)
ws['F1'].font = Font(bold=True)
ws['G1'].font = Font(bold=True)

#Задаём рамки вверх, вниз для заголовков
ws['A1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
ws['B1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
ws['C1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
ws['D1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
ws['E1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
ws['F1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))
ws['G1'].border=Border(top=Side(style='thick'),bottom=Side(style='thick'))

#Добавляем автофильтр
ws.auto_filter.ref = ws.dimensions

for result in one_phone:

    ws['A' + str(count)] = str(result['user']).replace('[]', '')
    ws['B' + str(count)] = str(result['dept']).replace('[]', '')
    ws['C' + str(count)] = str(result['tel']).replace('[]', '')
    ws['D' + str(count)] = str(result['login']).replace('[]', '')
    ws['E' + str(count)] = str(result['email']).replace('[]', '')
    ws['F' + str(count)] = str(result['location']).replace('[]', '')
    ws['G' + str(count)] = str(result['office']).replace('[]', '')

    count += 1

count = 1
for res_busy_numbers in busy_numbers:
    ws1['A' + str(count)] = res_busy_numbers
    count += 1

#Сохраняем файл
wb.save(f"Актуальные номера ({current_time.strftime('%d.%m.%Y')}).xlsx")
print('\nСкрипт выполнен...')










