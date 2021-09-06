from ldap3 import Server, Connection, SUBTREE, ALL, NTLM
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl import Workbook
from datetime import datetime
import re,operator,pandas as pd,getpass,time

#Функция для обновления даты и времени log.txt
def log_data_time():
    current_time = datetime.now()
    time_log = current_time.strftime("%d.%m.%Y %H:%M:%S")
    return time_log

#Логируем действие в файл
with open('log.txt', 'a') as f:
    print('-----------------------------------------', file=f)
    print(f"{log_data_time()} INFO: Скрипт запущен", file=f)

try:

    # Настройки подключения к AD
    server = Server('LDAP-DC.merlion.local', use_ssl=True, get_info=ALL)
    domain = 'MERLION\\'

    my_user = input('Введите пользователя: ')
    my_secret = getpass.getpass('Введите пароль: ')

    # Подключаемся к AD
    conn = Connection(server, user=domain + my_user, password=my_secret, authentication=NTLM,
                      return_empty_attributes=True, auto_bind=True)

    # Поисковый запрос
    entry_list = conn.extend.standard.paged_search(search_base='ou=users,ou=users_domain,dc=merlion,dc=local',
                                                   search_filter='(&(objectCategory=person)(objectClass=user) (!(userAccountControl:1.2.840.113556.1.4.803:=2)) (telephoneNumber=*) )',
                                                   search_scope=SUBTREE,
                                                   attributes=['extensionAttribute2', 'sAMAccountName', 'department','mail', 'telephoneNumber', 'Company','physicalDeliveryOfficeName', 'msRTCSIP-Line'],
                                                   paged_size=1000,
                                                   generator=False)

    # Закрываем соединение с AD
    conn.unbind()

    #Логируем действие в файл
    with open('log.txt', 'a') as f:
        print(f'{log_data_time()} INFO: Успешное подключение от пользователя: "{domain}{my_user}"', file=f)

    #Тайм-аут 2 сек.
    time.sleep(2)

    # Нормализуем данные
    for res_tel in entry_list:
        all_tel = str(res_tel['attributes']['telephoneNumber']) + ',' + str(res_tel['attributes']['msRTCSIP-Line'])
        clean_string = re.sub(r"[+()0-9 -]{10,20}", "", all_tel)
        iptel = re.findall(r'\d{4,5}', clean_string)
        # Получаем уникальные номера
        res_tel['iptel'] = set(iptel) if iptel else None

    # Собираем свой список из словарей
    persons_list = [
        dict(user=item['attributes']['extensionAttribute2'],
             dept=item['attributes']['department'],
             tel=item['iptel'],
             login=item['attributes']['sAMAccountName'],
             email=item['attributes']['mail'],
             location=item['attributes']['Company'],
             office=item['attributes']['physicalDeliveryOfficeName'],
             doubling=False,
             lost=False
             )

        for item in entry_list
        if item['iptel']
    ]

    # Логируем действие в файл
    with open('log.txt', 'a') as f:
        print(f"{log_data_time()} INFO: Нормализовал телефонные номера, создал необходимую структуры данных для продолжения работы", file=f)

    #Тайм-аут 2 сек.
    time.sleep(2)

    try:

        # Читаем файл "sp.txt"
        with open('sp.txt', 'r') as f:
            data_load_tel = f.read().splitlines()
            data_load_tel.sort()

        # Логируем действие в файл
        with open('log.txt', 'a') as f:
            print(f"{log_data_time()} INFO: Прочитал файл 'sp.txt'", file=f)

        # Тайм-аут 2 сек.
        time.sleep(2)

        # Один телефон одна запись
        one_phone = pd.DataFrame(persons_list).explode("tel").to_dict("records")

        # Добавляем данные в "data_list" для поиска дубликатов и пропущенных номеров
        data_list = []
        for telnumber in one_phone:
            data_list.append(telnumber['tel'])
            data_list_int = list(map(int, data_list))

        #Повторяющиеся и потерянные номера
        lost_data = list(map(str, list(set(range(min(data_list_int), max(data_list_int) + 1)) - set(data_list_int))))
        duplicates = [item for item in set(data_list) if data_list.count(item) > 1]

        # Логируем действие в файл
        with open('log.txt', 'a') as f:
            print(f"{log_data_time()} INFO: Форматируем данные к виду 'Один телефон, одна запись'", file=f)
            time.sleep(2)
            print(f"{log_data_time()} INFO: Находим повторяющиеся, потерянные номера", file=f)

        # Обновляем в список persons_list, меняем значение ключа repeat='yes'
        for repeat_res in one_phone:
            if repeat_res['tel'] in duplicates:
                repeat_res['doubling'] = True

        # Добавляем данные в словарь и обновляем данные для ключа "lost"
        for lost_res in lost_data:
            one_phone.append(
                dict(user=' ', dept=' ', tel=lost_res, login=' ', email=' ', location=' ', office=' ', doubling=False, lost=True))

        # Сортируем словарь по ключу 'tel'
        one_phone.sort(key=operator.itemgetter('tel'))

        # Тайм-аут 2 сек.
        time.sleep(2)

        # Логируем действие в файл
        with open('log.txt', 'a') as f:
            print(f"{log_data_time()} INFO: Изменил значения у повторяющихся, потерянных номеров с 'False' на 'True', отсортировал значения по телефонам'", file=f)

        # Создаем файл Excel
        wb = Workbook()
        ws = wb.active
        ws1 = wb.create_sheet()

        # Называем лист Excel
        ws.title = "Users DC=merlion, DC=local"
        ws1.title = "Занятые номера"

        # Заморозка первой строки, столбца
        wb["Users DC=merlion, DC=local"].freeze_panes = "B2"

        # Счётчик ячеек
        count = 2

        # Цвета для расскраски строк
        yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        magentaFill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')

        # Задаем заголовки
        ws['A1'] = 'Ф.И.О'
        ws['B1'] = 'Отдел'
        ws['C1'] = 'Внутренний телефон'
        ws['D1'] = 'Учётная запись'
        ws['E1'] = 'Электронная почта'
        ws['F1'] = 'Местоположение'
        ws['G1'] = 'Рабочее место'

        # Задаем ширину столбцов для первого листа
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 72
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 35
        ws.column_dimensions["F"].width = 57
        ws.column_dimensions["G"].width = 22

        # Задаем ширину столбцов для второго листа
        ws1.column_dimensions["A"].width = 60

        # Выравниваем заголовок по центру
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['B1'].alignment = Alignment(horizontal="center")
        ws['C1'].alignment = Alignment(horizontal="center")
        ws['D1'].alignment = Alignment(horizontal="center")
        ws['E1'].alignment = Alignment(horizontal="center")
        ws['F1'].alignment = Alignment(horizontal="center")
        ws['G1'].alignment = Alignment(horizontal="center")

        # Выделяем заголовок "жирным текстом"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)
        ws['D1'].font = Font(bold=True)
        ws['E1'].font = Font(bold=True)
        ws['F1'].font = Font(bold=True)
        ws['G1'].font = Font(bold=True)

        # Задаём рамки вверх, вниз для заголовков
        ws['A1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        ws['B1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        ws['C1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        ws['D1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        ws['E1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        ws['F1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))
        ws['G1'].border = Border(top=Side(style='thick'), bottom=Side(style='thick'))

        # Добавляем автофильтр
        ws.auto_filter.ref = ws.dimensions

        # Сохраняем данные на первый лист и красим
        for result in one_phone:
            if result['tel'] in data_load_tel:
                if result['lost'] == True:

                    ws['A' + str(count)] = ' '
                    ws['B' + str(count)] = ' '
                    ws['C' + str(count)] = result['tel']
                    ws['D' + str(count)] = ' '
                    ws['E' + str(count)] = ' '
                    ws['F' + str(count)] = ' '
                    ws['G' + str(count)] = ' '

                    # Красим желтым
                    ws['A' + str(count)].fill = yellowFill
                    ws['B' + str(count)].fill = yellowFill
                    ws['C' + str(count)].fill = yellowFill
                    ws['D' + str(count)].fill = yellowFill
                    ws['E' + str(count)].fill = yellowFill
                    ws['F' + str(count)].fill = yellowFill
                    ws['G' + str(count)].fill = yellowFill

                elif result['doubling'] == False:

                    ws['A' + str(count)] = str(result['user']).replace('[]', '')
                    ws['B' + str(count)] = str(result['dept']).replace('[]', '')
                    ws['C' + str(count)] = result['tel']
                    ws['D' + str(count)] = str(result['login']).replace('[]', '')
                    ws['E' + str(count)] = str(result['email']).replace('[]', '')
                    ws['F' + str(count)] = str(result['location']).replace('[]', '')
                    ws['G' + str(count)] = str(result['office']).replace('[]', '')

                elif result['doubling'] == True:

                    ws['A' + str(count)] = str(result['user']).replace('[]', '')
                    ws['B' + str(count)] = str(result['dept']).replace('[]', '')
                    ws['C' + str(count)] = result['tel']
                    ws['D' + str(count)] = str(result['login']).replace('[]', '')
                    ws['E' + str(count)] = str(result['email']).replace('[]', '')
                    ws['F' + str(count)] = str(result['location']).replace('[]', '')
                    ws['G' + str(count)] = str(result['office']).replace('[]', '')

                    # Красим пурпурным
                    ws['A' + str(count)].fill = magentaFill
                    ws['B' + str(count)].fill = magentaFill
                    ws['C' + str(count)].fill = magentaFill
                    ws['D' + str(count)].fill = magentaFill
                    ws['E' + str(count)].fill = magentaFill
                    ws['F' + str(count)].fill = magentaFill
                    ws['G' + str(count)].fill = magentaFill

                # Увеличиваем счётчик для номеров ячеек
                count += 1

        # Логируем действие в файл
        with open('log.txt', 'a') as f:
            print(f"{log_data_time()} INFO: Создал настройки для файла, заполнил данными лист 'Users DC=merlion, DC=local'", file=f)

        try:
            # Читаем файл "sp_net.txt"
            with open('sp_net.txt', 'r') as f:
                busy_numbers = f.read().splitlines()

                # Логируем действие в файл
                with open('log.txt', 'a') as f:
                    print(f"{log_data_time()} INFO: Файл 'sp_net.txt' прочитан", file=f)

                # Сохраняем "занятые телефоны" на второй лист
                count = 1
                for res_busy_numbers in busy_numbers:
                    ws1['A' + str(count)] = res_busy_numbers
                    count += 1

                # Логируем действие в файл
                with open('log.txt', 'a') as f:
                    print(f"{log_data_time()} INFO: Подготовил данные для второго листа 'Занятые номера'", file=f)

                try:

                    # Сохраняем файл
                    wb.save(f"Актуальные номера ({datetime.now().strftime('%d.%m.%Y')}).xlsx")

                    # Логируем действие в файл
                    with open('log.txt', 'a') as f:
                        print(f"{log_data_time()} INFO: Файл 'Актуальные номера ({datetime.now().strftime('%d.%m.%Y')}).xlsx сохранен'", file=f)

                        # Тайм-аут 2 сек.
                        time.sleep(2)
                        print(f"{log_data_time()} INFO: Скрипт отработал успешно!!!", file=f)

                except:
                    with open('log.txt', 'a') as f:
                        print(f"{log_data_time()} ERROR: Ошибка сохранения файла 'Актуальные номера ({datetime.now().strftime('%d.%m.%Y')}).xlsx'", file=f)

        except:

            # Логируем действие в файл
            with open('log.txt', 'a') as f:
                print(f"{log_data_time()} ERROR: Ошибка чтения файла 'sp_net.txt'", file=f)

    except:

        # Логируем действие в файл
        with open('log.txt', 'a') as f:
            print(f"{log_data_time()} ERROR: Не удалось считать данные из файла 'sp.txt'!!!", file=f)

except:
    # Логируем действие в файл
    with open('log.txt', 'a') as f:
        print(f"{log_data_time()} ERROR: Ошибка подключения, необходимо проверить имя пользователя или пароль!!!", file=f)


